import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from datetime import datetime

root = tk.Tk()
root.title("Laporan Keuangan Dafema")
root.geometry("600x400")

# Nama file Excel
FILE_EXCEL = "laporan_keuangan.xlsx"

def tambah_data():
    tanggal = entry_tanggal.get()
    deskripsi = entry_deskripsi.get()
    jumlah = entry_jumlah.get()
    try:
        # Validasi jumlah sebagai angka
        jumlah = float(jumlah)
        
        # Format nama bulan berdasarkan tanggal
        bulan = datetime.strptime(tanggal, "%Y-%m-%d").strftime("%B %Y")
        
        # Tambahkan data ke Excel
        tambah_data_ke_excel(bulan, tanggal, deskripsi, jumlah)

        # Tambahkan data ke tabel tkinter
        data.append({"Tanggal": tanggal, "Deskripsi": deskripsi, "Jumlah": jumlah})
        update_tabel()

        messagebox.showinfo("Sukses", "Data berhasil ditambahkan!")
    except ValueError:
        messagebox.showerror("Error", "Jumlah harus berupa angka!")

def tambah_data_ke_excel(bulan, tanggal, deskripsi, jumlah):
    try:
        # Jika file sudah ada, buka workbook
        workbook = load_workbook(FILE_EXCEL)
    except FileNotFoundError:
        # Jika file belum ada, buat workbook baru
        workbook = Workbook()
    
    # Jika tab (worksheet) bulan belum ada, buat baru
    if bulan not in workbook.sheetnames:
        workbook.create_sheet(title=bulan)
        # Tambahkan header jika worksheet masih kosong
        sheet = workbook[bulan]
        sheet.append(["Tanggal", "Deskripsi", "Jumlah"])
    else:
        sheet = workbook[bulan]
    
    # Tambahkan data ke worksheet
    sheet.append([tanggal, deskripsi, jumlah])
    
    # Simpan workbook
    workbook.save(FILE_EXCEL)

def update_tabel():
    # Bersihkan tabel tkinter
    for item in tree.get_children():
        tree.delete(item)
    # Masukkan data ke tabel tkinter
    for i, row in enumerate(data):
        tree.insert("", "end", values=(i+1, row["Tanggal"], row["Deskripsi"], row["Jumlah"]))

def ekspor_excel():
    # Hanya tampilkan pesan berhasil karena data sudah otomatis tersimpan ke Excel
    messagebox.showinfo("Sukses", f"Data berhasil diekspor ke file {FILE_EXCEL}!")

# Data lokal untuk tabel tkinter
data = []

# Frame untuk form input
frame_form = tk.Frame(root, padx=10, pady=10)
frame_form.pack(pady=10)

tk.Label(frame_form, text="Tanggal").grid(row=0, column=0, padx=5, pady=5)
entry_tanggal = DateEntry(frame_form, width=27, date_pattern="yyyy-MM-dd")
entry_tanggal.grid(row=0, column=1, padx=5, pady=5)

tk.Label(frame_form, text="Deskripsi").grid(row=1, column=0, padx=10, pady=10)
entry_deskripsi = tk.Entry(frame_form, width=30)
entry_deskripsi.grid(row=1, column=1, padx=10, pady=10)

tk.Label(frame_form, text="Jumlah").grid(row=2, column=0, padx=5, pady=5)
entry_jumlah = tk.Entry(frame_form, width=30)
entry_jumlah.grid(row=2, column=1, padx=5, pady=5)

tk.Button(frame_form, text="Tambah", command=tambah_data).grid(row=3, column=0, columnspan=2, pady=10)

# Tabel untuk menampilkan data
tree = ttk.Treeview(root, columns=("No", "Tanggal", "Deskripsi", "Jumlah"), show="headings")
tree.heading("No", text="No")
tree.heading("Tanggal", text="Tanggal")
tree.heading("Deskripsi", text="Deskripsi")
tree.heading("Jumlah", text="Jumlah")
tree.pack(pady=10)

# Tombol ekspor Excel
tk.Button(frame_form, text="Ekspor Excel", command=ekspor_excel).grid(row=4, column=0, columnspan=2, pady=10)

root.mainloop()
